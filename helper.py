import numpy as np 
import pandas as pd
from openpyxl import Workbook


##### 把dataframe的列名转为字典，键为列名，值为列的顺序（从1开始）
def col_names_to_dict(col_names):
    col_names_dict = {}
    for i, col_name in enumerate(col_names, 1):
        col_names_dict[col_name] = i
    return col_names_dict
# {('', ''): 1, ('', ''): 2}


##### 给每一个裁切方案加上组号，为了后续进行排序的时候把同一个方案的计划排到一起以及合并同一个方案的单元格
##### 补充相同的母材信息，把同组的纳期全部设定为组内订单纳期最靠前的日期
##### 后续按照生产计划号进行计划分组，具有同一生产计划号的计划安排挨在一起 
def fill_columns(df):
    df[('分组', '分组子列')] = None
    df[('新纳期', '新纳期子列')] = None
    df[('生产时间', '生产时间子列')] = None
    min_due_in_groups = {}
    ##### 给每一个裁切方案加上组号，为了后续进行排序的时候把同一个方案的计划排到一起以及合并同一个方案的单元格
    current_number = 1
    parent_column_names = [('母材信息', '钢卷号'), ('母材信息', '厚度(mm)'), ('母材信息', '宽度(mm)'), ('母材信息', '长度(mm)'), ('母材信息', '净重(kg)'), ('母材信息', '钢厂')]
    for index, row in df.iterrows():
        # 加工方式为空
        if pd.isna(row[('加工信息', '加工方式')]):
            df.at[index, ('分组', '分组子列')] = df.at[index - 1, ('分组', '分组子列')]
            df.at[index, ('加工信息', '加工方式')] = df.at[index - 1, ('加工信息', '加工方式')]
        else:
            df.at[index, ('分组', '分组子列')] = current_number
            current_number += 1

        # 母材信息为空
        if pd.isna(row[('母材信息', '钢卷号')]):
            for col_name in parent_column_names:
                df.at[index, col_name] = df.at[index - 1, col_name]
        
        # 加上生产准备时间
        if row[('加工信息', '加工方式')] != '/':
            df.at[index, ('生产时间', '生产时间子列')] = 19
        if row[('加工信息', '加工方式')] == '/':
            df.at[index, ('生产时间', '生产时间子列')] = 0

    for _, row in df.iterrows():
        group_num = row[('分组', '分组子列')]
        if group_num not in min_due_in_groups:
            min_due_in_groups[group_num] = row[('加工信息', '纳期')]
        else:
            if row[('加工信息', '纳期')] <= min_due_in_groups[group_num]:
                min_due_in_groups[group_num] = row[('加工信息', '纳期')]

    for index, row in df.iterrows():
        group_num = row[('分组', '分组子列')]
        df.at[index, ('新纳期', '新纳期子列')] = min_due_in_groups[group_num]
        

##### 读取dataframe，创建workbook和worksheet
def to_wb_ws(sorted_df):
    wb = Workbook()
    ws = wb.active
    for col_idx, (col_header, sub_header) in enumerate(sorted_df.columns):
        # 第一行和第二行为列名
        ws.cell(row = 1, column = col_idx + 1, value = col_header)
        ws.cell(row = 2, column = col_idx + 1, value = sub_header)

        # 从第三行开始输入一整列数据
        for row_idx, cell_value in enumerate(sorted_df[(col_header, sub_header)], start = 3):
            ws.cell(row = row_idx, column = col_idx + 1, value = cell_value)
    return wb,ws


##### key对应列名开始与结束的位置
def start_end_col_num(key, col_names_dict):
    set_start = False
    for i in col_names_dict:
        if i[0] == key:
            col_end_num = col_names_dict[i]
            if not set_start:
                col_start_num = col_names_dict[i]
                set_start = True
    return (col_start_num, col_end_num)


##### 合并第一行的列"母材信息"单元格和列"加工信息"单元格
def merge_column_names(ws, col_names_dict):
    keys = ['母材信息', '加工信息']
    for key in keys:
        start_column, end_column = start_end_col_num(key, col_names_dict)
        ws.merge_cells(start_row = 1, end_row = 1, start_column = start_column, end_column = end_column)


##### 合并所需合并的单元格(母材信息、加工方式、生产时间)
def combine_cells(sorted_df, ws, col_names_dict):
    for index, row in sorted_df.iterrows():
        if index == 0:
            group_start_row = index + 3
            group_end_row = index + 3 
            parent_start_row = index + 3
            parent_end_row = index + 3
            previous_row = row[('分组', '分组子列')]
            previous_parent_coil_number = row[('母材信息', '钢卷号')]
        else:
            if row[('分组', '分组子列')] == previous_row and row[('分组', '分组子列')] != '/':
                group_end_row  = index + 3
            else:
                if group_start_row != group_end_row:
                    ws.merge_cells(start_row = group_start_row, 
                                   start_column = col_names_dict[('生产时间', '生产时间子列')], 
                                   end_row = group_end_row, 
                                   end_column = col_names_dict[('生产时间', '生产时间子列')])
                
                    ws.merge_cells(start_row = group_start_row, 
                                   start_column = col_names_dict[('加工信息', '加工方式')], 
                                   end_row = group_end_row, 
                                   end_column = col_names_dict[('加工信息', '加工方式')])
                
                group_start_row, group_end_row = index + 3, index + 3 

            start_num, end_num = start_end_col_num('母材信息', col_names_dict)
            if row[('母材信息', '钢卷号')] == previous_parent_coil_number and row[('母材信息', '钢卷号')] != '/':
                parent_end_row = index + 3
            else:
                if parent_start_row != parent_end_row:
                    for i in range(start_num, end_num + 1):
                        ws.merge_cells(start_row = parent_start_row, end_row = parent_end_row, start_column = i, end_column = i)
                parent_start_row, parent_end_row = index + 3, index + 3

            previous_row = row[('分组', '分组子列')]
            previous_parent_coil_number = row[('母材信息', '钢卷号')]
    

##### 调整生产时间，如果相邻的计划不属于同一个方案但裁切方式相同，那么靠后的计划时间减去3分钟的换刀时间
def adjust_change_knife_time(sorted_df):
    for index, row in sorted_df.iterrows():
        if index == 0:
            previous_row_group = row[('分组', '分组子列')]
            previous_row_time = row[('生产时间', '生产时间子列')]
            previous_row_cut = row[('加工信息', '加工方式')]
        else:
            if row[('分组', '分组子列')] == previous_row_group:
                sorted_df.at[index, ('生产时间', '生产时间子列')] = previous_row_time
            else:
                if row[('加工信息', '加工方式')] == previous_row_cut and row[('加工信息', '加工方式')] != '/':
                    sorted_df.at[index, ('生产时间', '生产时间子列')] = row[('生产时间', '生产时间子列')] - 3

        # print(f"row[('生产时间', '生产时间子列')]: {row[('生产时间', '生产时间子列')]}")
        # print(f"previous_row_time: {previous_row_time}")
        
        previous_row_group = row[('分组', '分组子列')] 
        previous_row_time = row[('生产时间', '生产时间子列')]
        previous_row_cut = row[('加工信息', '加工方式')]


##### 计算总的生产时间
def compute_total_time(col_names_dict, ws):
    column_values = []
    column_index, _ = start_end_col_num('生产时间', col_names_dict) 
    for cell in ws.iter_rows(min_col = column_index, max_col = column_index, values_only = True):
        column_values.append(cell[0]) 
    column_values = column_values[2:]
    column_values = [i for i in column_values if i is not None]
    return sum(column_values)


##### 计算排刀外径（目前存在问题，内径不知道是否固定还是可以设置为508，密度也不知道是否是固定的）
def compute_outer_radius(total_weight, width, num, density = 7.85, inner_radius = 508):
    return np.sqrt(4000000 * total_weight / np.pi / width / num / density + inner_radius ** 2)
    # 单位是mm，需要注意需要转成m的情况


##### 检查排刀外径是否小于要求外径上限
def compare_outer_radius(coil_outer_radius, maximum_outer_radius):
    return (coil_outer_radius <= maximum_outer_radius)


##### 计算一卷母材有多长


##### 实现白晚班分配，实现不了，目前没有生产时间的计算规则，也没什么数据，根本没办法开始

