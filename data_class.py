import pandas as pd
from openpyxl import Workbook


class PlanData:
    def __init__(self, df):
        self.df = df
        self.fill_columns()
        self.get_sorted_df()


    @classmethod
    def read_data(cls, file_path):
        df = pd.read_excel(file_path, header = [0, 1])
        instance = PlanData(df)
        return instance
    

    def col_names_to_dict(self):
        col_names_dict = {}
        for i, col_name in enumerate(self.df.columns, 1):
            col_names_dict[col_name] = i
        return col_names_dict
    

    ##### 给每一个裁切方案加上组号，为了后续进行排序的时候把同一个方案的计划排到一起以及合并同一个方案的单元格
    ##### 补充相同的母材信息，把同组的纳期全部设定为组内订单纳期最靠前的日期
    def fill_columns(self):
        self.df[('分组', '分组子列')] = None
        self.df[('新纳期', '新纳期子列')] = None
        self.df[('生产时间', '生产时间子列')] = None
        min_due_in_groups = {}

        ##### 给每一个裁切方案加上组号，为了后续进行排序的时候把同一个方案的计划排到一起以及合并同一个方案的单元格
        current_number = 1
        parent_column_names = [('母材信息', '钢卷号'), ('母材信息', '厚度(mm)'), 
                               ('母材信息', '宽度(mm)'), ('母材信息', '长度(mm)'), 
                               ('母材信息', '净重(kg)'), ('母材信息', '钢厂')]
        for index, row in self.df.iterrows():
            if pd.isna(row[('加工信息', '加工方式')]):
                self.df.at[index, ('分组', '分组子列')] = self.df.at[index - 1, ('分组', '分组子列')]
                self.df.at[index, ('加工信息', '加工方式')] = self.df.at[index - 1, ('加工信息', '加工方式')]
            else:
                self.df.at[index, ('分组', '分组子列')] = current_number
                current_number += 1

            if pd.isna(row[('母材信息', '钢卷号')]):
                for col_name in parent_column_names:
                    self.df.at[index, col_name] = self.df.at[index - 1, col_name]
            
            ##### 先加上19分钟的生产准备时间
            if row[('加工信息', '加工方式')] != '/':
                self.df.at[index, ('生产时间', '生产时间子列')] = 19
            if row[('加工信息', '加工方式')] == '/':
                self.df.at[index, ('生产时间', '生产时间子列')] = 0

        for _, row in self.df.iterrows():
            group_num = row[('分组', '分组子列')]
            if group_num not in min_due_in_groups:
                min_due_in_groups[group_num] = row[('加工信息', '纳期')]
            else:
                if row[('加工信息', '纳期')] <= min_due_in_groups[group_num]:
                    min_due_in_groups[group_num] = row[('加工信息', '纳期')]

        for index, row in self.df.iterrows():
            group_num = row[('分组', '分组子列')]
            self.df.at[index, ('新纳期', '新纳期子列')] = min_due_in_groups[group_num]
    
        self.col_names_dict = self.col_names_to_dict()


    def get_sorted_df(self):
        self.sorted_df = self.df.sort_values(by = [('新纳期', '新纳期子列'), # 纳期为第一优先级
                                                   ('方案总刀数', '方案总刀数'),  # 刀数多的排前面进行生产
                                                   ('加工信息', '加工方式'), 
                                                   ('分组', '分组子列')], 
                                                   ascending=[True, False, True, True]).reset_index(drop = True)
        
    
    ##### 读取dataframe，创建workbook和worksheet
    def to_wb_ws(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        for col_idx, (col_header, sub_header) in enumerate(self.sorted_df.columns):
            self.ws.cell(row = 1, column = col_idx + 1, value = col_header)
            self.ws.cell(row = 2, column = col_idx + 1, value = sub_header)

            for row_idx, cell_value in enumerate(self.sorted_df[(col_header, sub_header)], start = 3):
                self.ws.cell(row = row_idx, column = col_idx + 1, value = cell_value)


    ##### 返回某一个连续出现的值的开始行数和结束行数，后续作为合并单元格的起始点
    @classmethod
    def start_end_col_num(cls, key, col_names_dict):
        set_start = False
        for i in col_names_dict:
            if i[0] == key:
                col_end_num = col_names_dict[i]
                if not set_start:
                    col_start_num = col_names_dict[i]
                    set_start = True
        return (col_start_num, col_end_num)


    ##### 合并所需合并的单元格(母材信息、加工方式、生产时间)
    def combine_cells(self):
        self.to_wb_ws()
        for index, row in self.sorted_df.iterrows():
            if index == 0:
                group_start_row = index + 3
                group_end_row = index + 3
                parent_start_row = index + 3
                parent_end_row = index + 3
                previous_row = row[('分组', '分组子列')]
                previous_parent_coil_number = row[('母材信息', '钢卷号')]
            else:
                if row[('分组', '分组子列')] == previous_row and row[('分组', '分组子列')] != '/':
                    group_end_row  = index + 3
                else:
                    if group_start_row != group_end_row:
                        self.ws.merge_cells(start_row = group_start_row, 
                                            start_column = self.col_names_dict[('生产时间', '生产时间子列')], 
                                            end_row = group_end_row, 
                                            end_column = self.col_names_dict[('生产时间', '生产时间子列')])
                    
                        self.ws.merge_cells(start_row = group_start_row, 
                                            start_column = self.col_names_dict[('加工信息', '加工方式')], 
                                            end_row = group_end_row, 
                                            end_column = self.col_names_dict[('加工信息', '加工方式')])
                        
                    group_start_row, group_end_row = index + 3, index + 3 
                
                start_num, end_num = self.start_end_col_num('母材信息', self.col_names_dict)
                if row[('母材信息', '钢卷号')] == previous_parent_coil_number and row[('母材信息', '钢卷号')] != '/':
                    parent_end_row = index + 3
                else:
                    if parent_start_row != parent_end_row:
                        for i in range(start_num, end_num + 1):
                            self.ws.merge_cells(start_row = parent_start_row, end_row = parent_end_row, start_column = i, end_column = i)
                    parent_start_row, parent_end_row = index + 3, index + 3

                previous_row = row[('分组', '分组子列')]
                previous_parent_coil_number = row[('母材信息', '钢卷号')]


    ##### 合并第一行的列"母材信息"单元格和列"加工信息"单元格
    def merge_column_names(self):
        keys = ['母材信息', '加工信息']
        for key in keys:
            start_column, end_column = self.start_end_col_num(key, self.col_names_dict)
            self.ws.merge_cells(start_row = 1, end_row = 1, start_column = start_column, end_column = end_column)            
    

    def output(self, file_path):
        self.wb.save(file_path)

